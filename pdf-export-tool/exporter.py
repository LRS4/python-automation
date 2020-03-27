""" 
Program:
- Reads Excel file lookup
- Creates folders (optional)
- Exports PDFs from PBI Desktop to folders
"""

import pyautogui
import time
import csv
import os
import psutil
import win32gui
from openpyxl import load_workbook

def main():

    w = win32gui

    # start clock
    start_time = time.time()

    # load workbook
    filename = r'C:\Users\lspencer\Lookup.xlsm'
    wb = load_workbook(filename = filename, data_only = True)

    # load worksheets
    lookup_sheet = wb['Lookup']

    # count rows
    row_count = lookup_sheet.max_row

    # find out screen size
    screen = pyautogui.size()

    # CREATE FOLDERS
    # loop through rows and extract folder names to set
    foldernames = set()
    for row in lookup_sheet.iter_rows(min_row=2, max_row=row_count):
        name = []
        for cell in row:
            name.append(cell.value)
        search_term = name[0]
        foldernames.add(search_term)

    # loop over set and create the unique folders
    for name in foldernames:
        newpath = rf'C:\Users\lspencer1\Reports\{name}'
        if not os.path.exists(newpath):
            os.makedirs(newpath)

    ## CREATE PDFS

    # initialise count
    count = 0

    # loop through providers row by row
    for row in lookup_sheet.iter_rows(min_row=2, max_row=row_count):
    
        # start clock for loop
        loop_start_time = time.time()

        # create blank array
        lookup_details = []

        # loop through row cells and add to array
        for cell in row:
            lookup_details.append(cell.value)

        # assign variables
        search_word = lookup_details[0]
        save_file_as = lookup_details[1]
        save_file_to = lookup_details[2]

        # check recursively if Power BI is current active window before proceeding
        def dashboardOpen():
            window_name = w.GetWindowText(w.GetForegroundWindow())
            if "Power BI Desktop" not in window_name:
                time.sleep(1)
                dashboardOpen()

        # fire function
        dashboardOpen()

        # click dropdown (if screen is desktop (1920x1080) else assume laptop resolution)
        if screen == (1920, 1080):
            pyautogui.moveTo(1518, 211)
        else:
            pyautogui.moveTo(2465, 404) 
        pyautogui.click()
        time.sleep(2)

        # click search bar
        if screen == (1920, 1080):
            pyautogui.moveTo(1482, 238) 
        else:
            pyautogui.moveTo(2367, 447)
        pyautogui.click()
        time.sleep(2)

        # select any text to overwrite
        pyautogui.hotkey('ctrl', 'a')

        # type search_word in to search bar
        pyautogui.typewrite(str(search_word))
        time.sleep(3)

        # click search_word on dropdown
        if screen == (1920, 1080):
            pyautogui.moveTo(1434, 266)
        else:
            pyautogui.moveTo(2376, 488)
        pyautogui.click()
        time.sleep(2)

        # click file
        if screen == (1920, 1080):
            pyautogui.moveTo(35, 38)
        else:
            pyautogui.moveTo(54, 68)
        pyautogui.click()
        time.sleep(2)

        # click export to pdf
        if screen == (1920, 1080):
            pyautogui.moveTo(91, 422)
        else:
            pyautogui.moveTo(186, 849)
        pyautogui.click()

        # wait for pdf to export
        time.sleep(10)

        # check recursively if acrobat is open yet
        def pdfOpen():
            window_name = w.GetWindowText(w.GetForegroundWindow())
            if "Adobe Acrobat" not in window_name:
                time.sleep(2)
                pdfOpen()
            
        # call function
        pdfOpen()

        # save file
        pyautogui.hotkey('ctrl', 'shift', 's')
        time.sleep(2)

        # enter file name
        pyautogui.typewrite(save_file_as)

        # enter desired folder into path
        pyautogui.hotkey('alt', 'd')
        pyautogui.typewrite(save_file_to)
        time.sleep(2)

        # confirm save
        pyautogui.press('enter')
        time.sleep(2)
        pyautogui.press('enter')
        time.sleep(2)
        pyautogui.press('enter')
        time.sleep(2)
        pyautogui.press('enter')
        
        # close pdf reader
        time.sleep(1)
        pyautogui.hotkey('alt', 'f4')
        time.sleep(2)
        pyautogui.click()

        # clear array for next iteration
        lookup_details = []

        # print iteration time and increment count
        count += 1
        print(f"PDF {count} for {search_term} exported in {time.time() - loop_start_time} seconds")

    # program finished
    pyautogui.alert(text=f'The program has finished. This program took {(time.time() - start_time) / 60} minutes to run. Have a nice day :)', title='Done!', button='OK')

if __name__ == "__main__":
    main()