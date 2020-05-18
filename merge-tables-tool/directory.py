# -*- coding: utf-8 -*-
import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook


class Directory():

    def __init__(self):
        self.directory = ""

    def get_table_name(self):
        """
        Gets name of table from user as input.

        Returns
        ---------
        Table name.

        """
        table = input("Name of table: ")
        return table

    def get_folder_from_user(self):
        """
        Gets folder path from user as input.

        Returns
        ---------
        Folder path.

        """
        folder = input("Enter the full path to the folder of excel files: ")
        return folder

    def get_folder_contents(self, directory):
        """
        Gets all excel files from the given directory.

        Returns
        -------
        Array of excel file names in the given directory.

        """
        self.directory = directory
        os.chdir(self.directory)

        files = []

        for file in os.listdir():
            files.append(file)

        return files

    def merge_all_tables(self, folder_contents, table_name):
        """
        Gets the given table name from each workbook and merges them 
        together into a single pandas DataFrame.

        Returns
        -------
        DataFrame of merged tables.

        """
        df = pd.DataFrame()
        for file in folder_contents:

            wb = load_workbook(filename=file, data_only=True)
            for ws in wb:
                for table in ws._tables:
                    if table.name == table_name:
                        data = ws[table.ref]

                        # Extract data from worksheet table
                        rows_list = []
                        for row in data:
                            cols = []
                            for col in row:
                                cols.append(col.value)
                            rows_list.append(cols)

                        # Create a pandas dataframe from the rows_list.
                        # The first row is the column names
                        table_df = pd.DataFrame(
                            data=rows_list[1:], index=None, columns=rows_list[0])
                        table_df.insert(0, 'Source', file,
                                        allow_duplicates=True)
                        df = df.append(table_df)

        return df
